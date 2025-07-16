from .web import Web
import re
from os.path import isfile
from .filemanager import FM
from .util import to_hash, json_serial
from openpyxl import load_workbook
from .tp import Link, IdTxt, IdTxtFk, CodTxt, CodTxtAgg
from functools import cached_property, cache
from datetime import datetime
import pandas as pd
import math
from enum import Enum
from typing import NamedTuple, Tuple, Dict, Any, List, Set
from collections import defaultdict
from types import MappingProxyType
from bs4 import Tag

import json


re_sp = re.compile(r"\s+")
re_hasta = re.compile(r"^[EX\d \+]+\(hasta 27/07/2007\)\s*\+", flags=re.IGNORECASE)

GRUPOS = MappingProxyType({
    "A1": ("A1", ),
    "A1A2": ("A1", "A2"),
    "A2": ("A2", ),
    "A2C1": ("A2", "C1"),
    "B": ("B", ),
    "C1": ("C1", ),
    "C1C2": ("C1", "C2"),
    "C2": ("C2", ),
    "C2E": ("C2", "E"),
    "E": ("E", ),
})


def get_codes(s: str):
    if s is None:
        return tuple()
    s = re_sp.sub(r" ", s).strip()
    s = re_hasta.sub(r"", s)
    if len(s) == 0:
        return tuple()
    s = re.sub(r"\b[0123]?\d/[01]?\d/20\d\d\b", "", s)
    s = re.sub(r"\s*/\s*", "/", s)
    cods: Set[str] = set()
    cods = cods.union(re.findall(r"\b(?:E|Z|H|C|EX|AC|EC|EB|FO|EJ|EK|EP)\d+", s))
    cods = cods.union(re.findall(r"\b(DLF|H\.V)\b", s))
    if re.search(r"AGRUPACION DE CUERPOS|INCLUYE CODIGOS", s):
        cods = cods.union(re.findall(r"\d+", s))
        z = s.split()[-1]
        if "/" in z:
            cods = cods.union(z.split("/"))
    for r in (r"\s*-\s*", r"\s*\+\s*"):
        spl = re.split(r, s)
        if len(spl) < 2:
            continue
        if all((" " not in x) for x in spl):
            cods = cods.union(spl)
    m = re.search(r"^(\d+) [oO] .*", s)
    if m is not None:
        cods.add(m.group(1))
    return tuple(sorted(cods, key=lambda x: (int(x) if x.isdigit() else -1, x)))


class ColType(NamedTuple):
    alias: Tuple[str, ...]
    type: str = None


class Col(Enum):
    ministerio_id = ColType(alias=('Minis.', ), type='Int64')
    ministerio_txt = ColType(alias=('Denominación Ministerio', ))
    centro_id = ColType(alias=('C.Dir', ), type='Int64')
    centro_txt = ColType(alias=('Denominación C.Dir', ))
    unidad_id = ColType(alias=('Unidad', ), type='Int64')
    unidad_txt = ColType(alias=('Denominación Unidad', ))
    unidad_pais_id = ColType(alias=('País U.', ), type='Int64')
    unidad_pais_txt = ColType(alias=('Denominación País U.', ))
    unidad_provincia_id = ColType(alias=('Provincia U.', ), type='Int64')
    unidad_provincia_txt = ColType(alias=('Denominación Provincia U.', ))
    unidad_localidad_id = ColType(alias=('Localidad U.', ), type='Int64')
    unidad_localidad_txt = ColType(alias=('Denominación Localidad U.', ))
    puesto_id = ColType(alias=('Puesto', ), type='Int64')
    puesto_txt = ColType(alias=('Denominación Larga', ))
    nivel = ColType(alias=('Nivel', ), type='Int64')
    especifico = ColType(alias=('C.Específ.', ))
    tipo = ColType(alias=('T.Pto.', ))
    provision = ColType(alias=('Provis.', ))
    administracion = ColType(alias=('Ad.Pu', ))
    grupo = ColType(alias=('Gr/Sb', ))
    cuerpo = ColType(alias=('Agr.cuer/cuer', ))
    titulacion = ColType(alias=('Tit.Académica', ))
    formacion = ColType(alias=('For.Espec.', ), type='Int64')
    pais_id = ColType(alias=('País', ), type='Int64')
    pais_txt = ColType(alias=('Denominación País', ))
    provincia_id = ColType(alias=('Provincia', ), type='Int64')
    provincia_txt = ColType(alias=('Denominación Provincia', ))
    localidad_id = ColType(alias=('Localidad', ), type='Int64')
    localidad_txt = ColType(alias=('Denominación Localidad', ))
    observaciones = ColType(alias=('Observaciones', ))
    estado = ColType(alias=('Estado', ))

    @classmethod
    def find_by_value(cls, v: str):
        v = re_sp.sub(" ", v).strip()
        for c in cls:
            if v in c.value.alias:
                return c

    def __str__(self):
        return self.name


class Clv(Enum):
    general = ('1.- GENERALES:', )
    unidades = ('2.- CODIGOS DE LAS UNIDADES:', )
    puestos = ('3.- CODIGOS DE LOS PUESTOS:', )
    tipo = ('4.- TIPO DE PUESTO', )
    provision = ('5.- PROVISION', )
    administracion = ('6.- ADSCRIPCION A ADMINISTRACION', )
    cuerpos = ('7.- ADSCRIPCION A CUERPOS', )
    titulaciones = ('8.- TITULACIONES ACADEMICAS', )
    formacion = ('9.- FORMACIÓN ESPECÍFICA', )
    observaciones = ('10.- OBSERVACIONES', )
    estados = ('11.- ESTADOS', )

    @classmethod
    def find_by_value(cls, v: str):
        v = re_sp.sub(" ", v).strip().upper()
        for c in cls:
            if v in c.value:
                return c

    def __str__(self):
        return self.name


class Unidad(NamedTuple):
    id: int
    txt: str
    centro: int
    localidad: int
    provincia: int


class Puesto(NamedTuple):
    id: int
    txt: str
    unidad: int
    localidad: int
    provincia: int
    nivel: int
    especifico: float
    tipo: str
    provision: str
    administracion: str
    grupos: Tuple[str, ...]
    formacion: str
    vacante: bool
    titulaciones: Tuple[str, ...]
    cuerpos: Tuple[str, ...]
    observaciones: Tuple[str, ...]


def _parse_if_str(obj: Any) -> Any:
    if not isinstance(obj, str):
        return obj
    fake_bl = "[~~~~~NNNN~~~~~]"
    val: str = obj
    val = val.replace(r"\n", fake_bl)
    val = re_sp.sub(" ", val)
    val = val.replace(fake_bl, r"\n")
    val = val.strip()
    val = re.sub(r'(["\'])\s*\.\s*$', r'\1', val)
    val = re.sub(r'^\s*(["\'])\s*', r'\1', val)
    while len(val) > 1 and val[0] == val[-1] and val[0] in ("'", '"'):
        val = val[1:-1].strip()
    if len(val) == 0:
        return None
    val = re.sub(r"\s*\(\s*", " (", val)
    val = re.sub(r"\s*\)\s*", ") ", val)
    val = re.sub(r"\s*\[\s*", " [", val)
    val = re.sub(r"\s*\]\s*", "] ", val)
    val = val.strip()
    return val


def _parse(val):
    val = _parse_if_str(val)
    if isinstance(val, (int, float)) and math.isnan(val):
        return None
    if isinstance(val, float):
        i = int(val)
        if i == val:
            return i
    return val


class RPTError(ValueError):
    def __init__(self, link: Link, msg: str):
        super().__init__(link.href+" "+msg)


class RPTFinder:
    ROOT = "https://transparencia.gob.es/transparencia/transparencia_Home/index/PublicidadActiva/OrganizacionYEmpleo/Relaciones-Puestos-Trabajo.html"

    @cache
    def get(self):
        return Rpt(
            via=RPTFinder.ROOT,
            links=self.get_links()
        )

    @cache
    def get_links(self):
        arr: list[Rpt] = list()
        s = Web()
        s.get(RPTFinder.ROOT)
        a = s.soup.select_one("article li a[title*='funcionario'][href*='.xlsx']")
        text = re_sp.sub(" ", a.get_text()).strip()
        link = a.attrs["href"]
        arr.append(Link(href=link, text=text))
        links = [a.attrs["href"] for a in s.soup.select("div.title-item-div a[href]")]
        for link in links:
            s.get(link)
            for a in s.soup.findAll("a", string=re.compile(r".*\bfuncionarios?\b.*", flags=re.IGNORECASE)):
                li: Tag = a.find_parent("li")
                a: Tag = li.find("a", href=re.compile(r"\.xlsx$"))
                if a is not None:
                    text = re_sp.sub(" ", a.get_text()).strip()
                    lk = Link(href=a.attrs["href"], text=text)
                    if lk not in arr:
                        arr.append(lk)
        return tuple(arr)


class Rpt:
    def __init__(self, links: Tuple[Link], via: str):
        self.__links = links
        self.__via = via

    @cache
    def __get_local_path(self, link: Link) -> str:
        file = FM.resolve_path(f"dwn/rpt_{to_hash(link.href)}.xlsx")
        if not isfile(file):
            r = Web().s.get(link.href, verify=False)
            FM.dump(file, r.content)
        return file

    @property
    def link(self):
        return self.__links[0]

    @property
    def __file(self):
        return self.__get_local_path(self.__links[0])

    @property
    def via(self):
        return self.__via

    @cached_property
    def grupos(self):
        grupos: Set[str] = set()
        for vls in GRUPOS.values():
            grupos = grupos.union(vls)
        return tuple(sorted(grupos))

    @cached_property
    def fecha(self):
        wb = load_workbook(self.__file)
        md = wb.properties.modified
        if not isinstance(md, datetime):
            raise RPTError(self.link, "modified date not found")
        return md

    @cache
    def __get_df(self):
        df = pd.read_excel(self.__file, skiprows=3)
        if df is None or df.empty:
            raise ValueError(self.link.href+" can't be read")
        rename: Dict[str, str] = {}
        retype: Dict[str, str] = {}
        for name in map(str, df.columns):
            col = Col.find_by_value(name)
            if col is not None:
                rename[name] = col.name
                if col.value.type is not None:
                    retype[col.name] = col.value.type
        df = df[list(rename.keys())].rename(columns=rename)
        df = df.map(_parse)
        df = df.dropna(how='all', axis=0)
        df = df.dropna(how='all', axis=1)
        missing = set(c.name for c in Col).difference(map(str, df.columns))
        if len(missing) > 0:
            raise RPTError(self.link, "missing columns: " + ", ".join(sorted(missing)))
        for col, tp in retype.items():
            df[col] = df[col].astype(tp)
        for (pro_id, pro_txt, pais_id, pais_txt) in (
            map(str, (Col.provincia_id, Col.provincia_txt, Col.pais_id, Col.pais_txt)),
            map(str, (Col.unidad_provincia_id, Col.unidad_provincia_txt, Col.unidad_pais_id, Col.unidad_pais_txt)),
        ):
            isKO = (df[pro_id] == 60) & (df[pro_txt] != 'EXTRANJERO')
            if isKO.any():
                raise RPTError(self.link, f"{pro_id}==60 && {pro_txt}!=EXTRANJERO")
            df.loc[df[pro_id] == 60, pro_txt] = df[pais_txt]
            df.loc[df[pro_id] == 60, pro_id] = -df[pais_id]

        c_id, c_txt, u_id, u_txt = map(str, (Col.centro_id, Col.centro_txt, Col.unidad_id, Col.unidad_txt))
        cEmpty = (df[c_id].isnull() & df[c_txt].isnull() & ~df[u_id].isnull())
        df.loc[cEmpty, c_id] = -df[u_id]
        df.loc[cEmpty, c_txt] = df[u_txt]

        geoCl = list(map(str, (Col.pais_id, Col.provincia_id, Col.localidad_id)))
        geoOK = (df[geoCl].isnull().all(axis=1) | df[geoCl].notnull().all(axis=1)).all()
        if not geoOK:
            raise RPTError(self.link, f"<{Col.pais_id}, {Col.provincia_id}, {Col.localidad_id}> inconsistent")

        def __check_vals(col: Col, *vals, null=False):
            c = df[col.name]
            isOk = (None in vals or c.notna().all()) and c.isin(list(vals)).all()
            if isOk:
                return True
            real = tuple(c.sort_values().unique().tolist())
            raise RPTError(self.link, f"{col.name} is {real} instead not {vals}")

        __check_vals(Col.estado, "V", "NV")
        __check_vals(Col.grupo, None, *GRUPOS.keys())

        return df

    @cache
    def __get_claves(self):
        data: Dict[Clv, Set[CodTxt]] = defaultdict(set)
        for link in self.__links:
            clv = None
            for line in self.__iter_claves(link):
                aux = Clv.find_by_value(line)
                if aux is not None:
                    clv = aux
                    continue
                if clv is None:
                    raise RPTError(self.link, "error in 2º sheet")
                c, t = line.split(None, 1)
                txt = _parse(t)
                txt = re_hasta.sub("", txt)
                data[clv].add(CodTxt(cod=c, txt=txt))

        missing = set(c for c in Clv).difference(data.keys())
        if len(missing) > 0:
            raise RPTError(self.link, "missing in 2º sheet: " + ", ".join(sorted(map(str, missing))))
        clvs = {k: tuple(sorted(v)) for k, v in data.items()}
        return MappingProxyType(clvs)

    def __iter_claves(self, link: Link):
        local = self.__get_local_path(link)
        wb = load_workbook(local)
        sh = wb.worksheets[1]
        flag = True
        for row in sh.iter_rows(values_only=True):
            row = list(map(_parse, row))
            while len(row) > 0 and row[-1] is None:
                row.pop()
            if len(row) == 0:
                continue
            line = row[0]
            if not isinstance(line, str) or len(row) > 1:
                raise RPTError(self.link, "error in 2º sheet")
            if flag:
                if line.upper() == "CLAVES UTILIZADAS":
                    flag = False
                    continue
                raise ValueError(self.link, "error in 2º sheet")
            if len(line.split()) < 2:
                raise ValueError(self.link, "error in 2º sheet")
            yield line

    def __get_uniq(self, *args: str):
        cols = list(args)
        df = self.__get_df()[cols]
        df = df.drop_duplicates()
        df = df.dropna(how='all', axis=0)
        df = df.sort_values(by=cols)
        return df

    def __iter_idtxt(self, idCol: Col, txtCol: Col):
        idName = idCol.name
        txtName = txtCol.name
        for row in self.__get_uniq(idName, txtName).to_dict(orient="records"):
            yield IdTxt(id=row[idName], txt=row[txtName])

    def __iter_idtxt_fk(self, idCol: Col, txtCol: Col, fkCol: Col):
        idName = idCol.name
        txtName = txtCol.name
        fkName = fkCol.name
        for row in self.__get_uniq(idName, txtName, fkName).to_dict(orient="records"):
            yield IdTxtFk(id=row[idName], txt=row[txtName], fk=row[fkName])

    @cached_property
    def ministerios(self):
        return tuple(self.__iter_idtxt(Col.ministerio_id, Col.ministerio_txt))

    @cached_property
    def centros(self):
        return tuple(self.__iter_idtxt_fk(Col.centro_id, Col.centro_txt, Col.ministerio_id))

    @cached_property
    def unidades(self):
        arr: List[Unidad] = []
        u_id, u_txt, c_id, l_id, p_id = map(str, (Col.unidad_id, Col.unidad_txt, Col.centro_id, Col.unidad_localidad_id, Col.unidad_provincia_id))
        for row in self.__get_uniq(u_id, u_txt, c_id, l_id, p_id).to_dict(orient="records"):
            arr.append(Unidad(
                id=row[u_id],
                txt=row[u_txt],
                centro=row[c_id],
                localidad=row[l_id],
                provincia=row[p_id]
            ))
        return tuple(arr)

    @cached_property
    def pais(self):
        aux1 = self.__iter_idtxt(Col.pais_id, Col.pais_txt)
        aux2 = self.__iter_idtxt(Col.unidad_pais_id, Col.unidad_pais_txt)
        return tuple(sorted(set(aux1).union(aux2)))

    @cached_property
    def provincias(self):
        aux1 = self.__iter_idtxt_fk(Col.provincia_id, Col.provincia_txt, Col.pais_id)
        aux2 = self.__iter_idtxt_fk(Col.unidad_provincia_id, Col.unidad_provincia_txt, Col.unidad_pais_id)
        return tuple(sorted(set(aux1).union(aux2)))

    @cached_property
    def localidades(self):
        aux1 = self.__iter_idtxt_fk(Col.localidad_id, Col.localidad_txt, Col.provincia_id)
        aux2 = self.__iter_idtxt_fk(Col.unidad_localidad_id, Col.unidad_localidad_txt, Col.unidad_provincia_id)
        return tuple(sorted(set(aux1).union(aux2)))

    def __get_clave(self, clv: Clv, needed: Set[str]):
        agg: Dict[str, Set[str]] = defaultdict(set)
        arr = list(self.__get_claves()[clv])
        done = set(a.cod for a in arr)
        for c in list(arr):
            for a in get_codes(c.txt):
                agg[c.cod].add(a)
                if a not in done:
                    arr.append(CodTxt(cod=a, txt='¿?'))
                    done.add(a)
        for a in needed:
            if a not in done:
                arr.append(CodTxt(cod=a, txt='¿?'))
                done.add(a)

        def __toCodTxtAgg(c: CodTxt):
            aux = tuple(sorted(agg[c.cod]))
            return CodTxtAgg(cod=c.cod, txt=c.txt, agg=aux)

        return tuple(map(__toCodTxtAgg, arr))

    @cached_property
    def tipos(self):
        return self.__get_claves()[Clv.tipo]

    @cached_property
    def provision(self):
        return self.__get_claves()[Clv.provision]

    @cached_property
    def adminsitracion(self):
        return self.__get_claves()[Clv.administracion]

    @cached_property
    def formaciones(self):
        return self.__get_claves()[Clv.formacion]

    @cached_property
    def cuerpos(self):
        needed = {t for p in self.puestos for t in p.cuerpos}
        return self.__get_clave(Clv.cuerpos, needed)

    @cached_property
    def titulaciones(self):
        needed = {t for p in self.puestos for t in p.titulaciones}
        return self.__get_clave(Clv.titulaciones, needed)

    @cached_property
    def observaciones(self):
        needed = {t for p in self.puestos for t in p.observaciones}
        return self.__get_clave(Clv.observaciones, needed)

    @cached_property
    def puestos(self):
        def __split(s: str):
            if s is None:
                return tuple()
            return tuple(re.split(r"\s*,\s*", s))

        arr: List[Puesto] = []
        cols = tuple(map(str, (
            Col.puesto_id, Col.puesto_txt, Col.unidad_id, Col.nivel, Col.especifico,
            Col.tipo, Col.provision, Col.administracion, Col.grupo, Col.cuerpo,
            Col.titulacion, Col.formacion, Col.observaciones,
            Col.estado, Col.localidad_id, Col.unidad_localidad_id, Col.provincia_id, Col.unidad_provincia_id
        )))
        for row in self.__get_uniq(*cols).to_dict(orient="records"):
            arr.append(Puesto(
                id=row[Col.puesto_id.name],
                txt=row[Col.puesto_txt.name],
                unidad=row[Col.unidad_id.name],
                localidad=row[Col.localidad_id.name] or row[Col.unidad_localidad_id.name],
                provincia=row[Col.provincia_id.name] or row[Col.unidad_provincia_id.name],
                grupos=GRUPOS.get(row[Col.grupo.name], tuple()),
                nivel=row[Col.nivel.name],
                especifico=row[Col.especifico.name],
                tipo=row[Col.tipo.name],
                provision=row[Col.provision.name],
                administracion=row[Col.administracion.name],
                formacion=row[Col.formacion.name],
                vacante=row[Col.estado.name] == "V",
                titulaciones=__split(row[Col.titulacion.name]),
                cuerpos=__split(row[Col.cuerpo.name]),
                observaciones=__split(row[Col.observaciones.name]),
            ))
        return tuple(arr)

    def complete(self, p: Puesto):
        def __complete(vals: Tuple[str, ...], data: Dict[str, CodTxtAgg]):
            size = 0
            arr = list(vals)
            while size != len(arr):
                size = len(arr)
                for a in tuple(arr):
                    for x in data[a].agg:
                        if x not in arr:
                            arr.append(x)
            return tuple(arr)

        p = p._replace(
            cuerpos=__complete(p.cuerpos, {c.cod: c for c in self.cuerpos}),
            titulaciones=__complete(p.titulaciones, {c.cod: c for c in self.titulaciones}),
            observaciones=__complete(p.observaciones, {c.cod: c for c in self.observaciones}),
        )
        return p


if __name__ == '__main__':
    r = RPTFinder()
    r = r.get()

    print(json.dumps(json_serial(r.tipos), indent=2, default=json_serial))
